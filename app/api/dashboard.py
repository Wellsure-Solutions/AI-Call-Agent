from __future__ import annotations

import csv, io, json
from datetime import datetime, timezone
from typing import Any
from uuid import uuid4
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import Response
from openpyxl import load_workbook
try:
    import xlrd
except ImportError:  # pragma: no cover
    xlrd = None
from pydantic import BaseModel, Field

from app.analytics.dashboard import build_dashboard
from app.export.exporters import csv_bytes, rows, xlsx_bytes
from app.storage.json_storage import JsonStorage

router = APIRouter(prefix="/api", tags=["dashboard"])
storage = JsonStorage()

class CampaignIn(BaseModel):
    name: str
    description: str = ""
    status: str = "active"

class SettingsIn(BaseModel):
    default_ai_prompt: str = ""
    voice: str = ""
    language: str = "en"
    model: str = "gpt-4.1-mini"
    temperature: float = 0.7
    max_call_duration: int = 600
    retry_count: int = 1
    business_prompt_template: str = "You are calling {business_name}. They are a {category} business. Use their business name naturally."
    storage_location: str = "data/"
    export_settings: dict[str, Any] = Field(default_factory=dict)
    future_database_type: str = "postgresql"

@router.get("/dashboard")
def dashboard():
    return build_dashboard(storage.list("campaigns"), storage.list("leads"), storage.list("calls"), storage.list("responses"))

@router.get("/campaigns")
def campaigns(): return storage.list("campaigns")

@router.post("/campaigns")
def create_campaign(data: CampaignIn):
    now = datetime.now(timezone.utc).isoformat()
    return storage.create("campaigns", {**data.model_dump(), "created_at": now, "archived": data.status == "archived"})

@router.put("/campaigns/{campaign_id}")
def update_campaign(campaign_id: str, data: CampaignIn):
    item = storage.update("campaigns", campaign_id, data.model_dump())
    if not item: raise HTTPException(404, "Campaign not found")
    return item

@router.post("/campaigns/{campaign_id}/duplicate")
def duplicate_campaign(campaign_id: str):
    item = storage.get("campaigns", campaign_id)
    if not item: raise HTTPException(404, "Campaign not found")
    clone = {**item, "id": str(uuid4()), "name": f"{item.get('name')} Copy", "created_at": datetime.now(timezone.utc).isoformat()}
    return storage.create("campaigns", clone)

@router.delete("/campaigns/{campaign_id}")
def delete_campaign(campaign_id: str):
    return {"deleted": storage.delete("campaigns", campaign_id)}

@router.get("/leads")
def leads(campaign_id: str | None = None):
    data = storage.list("leads")
    return [l for l in data if not campaign_id or l.get("campaign_id") == campaign_id]

@router.post("/leads/upload")
async def upload_leads(campaign_id: str = Form(...), mapping: str = Form("{}"), file: UploadFile | None = File(None), pasted: str = Form("")):
    maps = json.loads(mapping or "{}")
    raw = []
    if file:
        content = await file.read()
        name = (file.filename or "").lower()
        if name.endswith(".csv"):
            raw = list(csv.DictReader(io.StringIO(content.decode("utf-8-sig"))))
        elif name.endswith(".xlsx"):
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active; rows_iter = list(ws.iter_rows(values_only=True)); wb.close()
            headers = [str(v or "").strip() for v in rows_iter[0]] if rows_iter else []
            raw = [dict(zip(headers, row)) for row in rows_iter[1:]]
        elif name.endswith(".xls"):
            if xlrd is None:
                raise HTTPException(400, "xlrd is required for XLS imports")
            book = xlrd.open_workbook(file_contents=content)
            sheet = book.sheet_by_index(0)
            headers = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)] if sheet.nrows else []
            raw = [dict(zip(headers, sheet.row_values(row))) for row in range(1, sheet.nrows)]
        else: raise HTTPException(400, "Upload CSV, XLS, or XLSX")
    elif pasted.strip():
        raw = list(csv.DictReader(io.StringIO(pasted), delimiter="\t"))
    else: raise HTTPException(400, "Provide a file or pasted spreadsheet data")
    imported, errors = [], []
    for i, row in enumerate(raw, 2):
        lead = {field: row.get(src) for field, src in maps.items() if src}
        if not maps:
            lead = {k.lower().strip().replace(" ", "_"): v for k, v in row.items()}
        lead.setdefault("business_name", lead.get("business") or lead.get("company"))
        lead.setdefault("phone", lead.get("phone_number"))
        missing = [f for f in ("business_name", "phone", "category") if not lead.get(f)]
        if missing: errors.append({"row": i, "error": f"Missing {', '.join(missing)}"}); continue
        lead.update({"campaign_id": campaign_id, "status": "pending", "created_at": datetime.now(timezone.utc).isoformat(), "custom_fields": {k: v for k, v in row.items() if k not in maps.values()}})
        imported.append(storage.create("leads", lead))
    return {"imported": len(imported), "errors": errors, "leads": imported[:50]}

@router.get("/calls")
def calls(q: str | None = None, campaign_id: str | None = None, status: str | None = None, category: str | None = None):
    data = storage.list("calls")
    def ok(c):
        blob = " ".join(str(c.get(k, "")) for k in ("business_name", "phone", "category", "campaign", "summary", "transcript")).lower()
        return (not q or q.lower() in blob) and (not campaign_id or c.get("campaign_id") == campaign_id) and (not status or c.get("status") == status) and (not category or c.get("category") == category)
    return [c for c in data if ok(c)]

@router.get("/calls/{call_id}")
def call_detail(call_id: str):
    call = storage.get("calls", call_id)
    if not call: raise HTTPException(404, "Call not found")
    response = next((r for r in storage.list("responses") if r.get("call_id") == call_id), {})
    return {"call": call, "response": response}

@router.get("/analytics")
def analytics(): return dashboard()

@router.get("/settings")
def get_settings(): return storage.settings()

@router.put("/settings")
def put_settings(settings: SettingsIn): return storage.save_settings(settings.model_dump())

@router.get("/export/{kind}")
def export(kind: str):
    data_rows = rows(storage.list("calls"), storage.list("responses"), storage.list("campaigns"))
    if kind == "csv": return Response(csv_bytes(data_rows), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=calls.csv"})
    if kind in {"excel", "xlsx"}: return Response(xlsx_bytes(data_rows), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=calls.xlsx"})
    if kind == "json": return {"campaigns": storage.list("campaigns"), "leads": storage.list("leads"), "calls": storage.list("calls"), "responses": storage.list("responses")}
    raise HTTPException(404, "Unknown export format")
