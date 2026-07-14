from __future__ import annotations

import csv
import json
import re
from datetime import datetime, timezone
from io import BytesIO, StringIO
from pathlib import Path
from threading import Lock
from typing import Any
from uuid import uuid4

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover
    Workbook = None
    load_workbook = None

from app.core.models import CallSession

REQUIRED_LEAD_FIELDS = ("business_name", "phone_number", "category")
EXPORT_HEADERS = [
    "call_id",
    "lead_id",
    "business_name",
    "phone",
    "category",
    "notes",
    "call_status",
    "interested",
    "callback_requested",
    "duration",
    "timestamp",
    "summary",
    "transcript",
    "structured_response",
]


class JsonCallStore:
    """Local JSON persistence for leads and call results.

    Files are intentionally simple and isolated so this class can later be
    swapped for a database-backed repository without changing API handlers.
    """

    def __init__(self, data_dir: Path) -> None:
        self.data_dir = data_dir
        self.results_dir = data_dir / "call_results"
        self.leads_path = data_dir / "leads.json"
        self._lock = Lock()

    def append_call(self, session: CallSession) -> dict[str, Any]:
        self.results_dir.mkdir(parents=True, exist_ok=True)
        record = self._session_to_record(session)
        with self._lock:
            path = self.results_dir / f"{record['call_id']}.json"
            path.write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
        return record

    def list_calls(self) -> list[dict[str, Any]]:
        self.results_dir.mkdir(parents=True, exist_ok=True)
        calls: list[dict[str, Any]] = []
        for path in sorted(self.results_dir.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True):
            try:
                calls.append(json.loads(path.read_text(encoding="utf-8")))
            except json.JSONDecodeError:
                continue
        return calls

    def get_call(self, call_id: str) -> dict[str, Any] | None:
        path = self.results_dir / f"{call_id}.json"
        if not path.exists():
            return None
        return json.loads(path.read_text(encoding="utf-8"))

    def list_leads(self) -> list[dict[str, Any]]:
        if not self.leads_path.exists():
            return []
        try:
            return json.loads(self.leads_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return []


    def get_lead(self, lead_id: str) -> dict[str, Any] | None:
        for lead in self.list_leads():
            if lead.get("lead_id") == lead_id:
                return lead
        return None

    def update_lead(self, lead_id: str, **updates: Any) -> dict[str, Any] | None:
        leads = self.list_leads()
        updated: dict[str, Any] | None = None
        for lead in leads:
            if lead.get("lead_id") == lead_id:
                lead.update({key: value for key, value in updates.items() if value is not None})
                lead["updated_at"] = datetime.now(timezone.utc).isoformat()
                updated = lead
                break
        if updated is None:
            return None
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self.leads_path.write_text(json.dumps(leads, ensure_ascii=False, indent=2), encoding="utf-8")
        return updated

    def import_leads(self, rows: list[dict[str, Any]]) -> dict[str, Any]:
        """Validate and import normalized lead rows.

        Returns an enterprise-friendly summary so UI/API callers can show what
        happened instead of silently dropping bad rows.
        """
        now = datetime.now(timezone.utc).isoformat()
        leads = self.list_leads()
        existing = {self._phone_key(lead.get("phone_number", "")) for lead in leads}
        imported: list[dict[str, Any]] = []
        rejected: list[dict[str, Any]] = []
        duplicate_count = 0

        for index, row in enumerate(rows, start=1):
            normalized, errors = self.validate_lead(row)
            if errors:
                rejected.append({"row": index, "errors": errors, "lead": row})
                continue
            phone_key = self._phone_key(normalized["phone_number"])
            if phone_key in existing:
                duplicate_count += 1
                rejected.append({"row": index, "errors": ["Duplicate phone number"], "lead": normalized})
                continue
            existing.add(phone_key)
            lead = {
                "lead_id": str(uuid4()),
                "business_name": normalized["business_name"],
                "phone_number": normalized["phone_number"],
                "category": normalized["category"],
                "notes": normalized.get("notes", ""),
                "created_at": now,
                "status": "new",
            }
            leads.append(lead)
            imported.append(lead)

        self.data_dir.mkdir(parents=True, exist_ok=True)
        self.leads_path.write_text(json.dumps(leads, ensure_ascii=False, indent=2), encoding="utf-8")
        return {
            "imported": len(imported),
            "duplicates": duplicate_count,
            "rejected": rejected,
            "total_submitted": len(rows),
            "leads": imported,
        }

    def validate_lead(self, row: dict[str, Any]) -> tuple[dict[str, str], list[str]]:
        normalized = {
            "business_name": self._clean_text(row.get("business_name")),
            "phone_number": self._normalize_phone(row.get("phone_number")),
            "category": self._clean_text(row.get("category")),
            "notes": self._clean_text(row.get("notes")),
        }
        errors: list[str] = []
        for field in REQUIRED_LEAD_FIELDS:
            if not normalized[field]:
                errors.append(f"Missing {field.replace('_', ' ')}")
        digit_count = len(re.sub(r"\D", "", normalized["phone_number"]))
        if normalized["phone_number"] and digit_count < 7:
            errors.append("Phone number is too short")
        return normalized, errors

    def parse_upload(self, content: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]]]:
        suffix = Path(filename).suffix.lower()
        if suffix in {".xlsx", ".xls"}:
            if load_workbook is None:
                raise RuntimeError("openpyxl is required for Excel uploads.")
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            workbook.close()
            if not rows:
                return [], []
            headers = [str(value or "").strip() for value in rows[0]]
            return headers, [dict(zip(headers, values)) for values in rows[1:] if any(values)]
        text = content.decode("utf-8-sig")
        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        except csv.Error:
            dialect = csv.excel_tab if "\t" in sample else csv.excel
        reader = csv.DictReader(StringIO(text), dialect=dialect)
        headers = [header.strip() for header in (reader.fieldnames or [])]
        rows: list[dict[str, Any]] = []
        for row in reader:
            rows.append({str(key or "").strip(): value for key, value in row.items()})
        return headers, rows

    def export_calls(self, fmt: str) -> tuple[str, bytes, str]:
        calls = self.list_calls()
        if fmt == "json":
            return "call_results.json", json.dumps(calls, ensure_ascii=False, indent=2).encode("utf-8"), "application/json"
        rows = [self._flat_call(call) for call in calls]
        headers = list(rows[0].keys()) if rows else EXPORT_HEADERS
        if fmt == "csv":
            out = StringIO()
            writer = csv.DictWriter(out, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)
            return "call_results.csv", out.getvalue().encode("utf-8"), "text/csv"
        if Workbook is None:
            raise RuntimeError("openpyxl is required for Excel export.")
        wb = Workbook()
        ws = wb.active
        ws.title = "Call Results"
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header, "") for header in headers])
        bio = BytesIO()
        wb.save(bio)
        wb.close()
        return "call_results.xlsx", bio.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def _session_to_record(self, session: CallSession) -> dict[str, Any]:
        meta = session.metadata or {}
        answers = session.answers or {}
        interest = answers.get("interested", answers.get("interest_level", "unknown"))
        callback = answers.get("callback_approved", "unknown")
        return {
            "call_id": session.call_id,
            "lead_id": meta.get("lead_id"),
            "business_name": meta.get("business_name", ""),
            "phone": session.phone_number or meta.get("phone_number", ""),
            "category": meta.get("category", ""),
            "notes": meta.get("notes", ""),
            "call_status": session.status,
            "interested": interest in {"hot", "warm", "yes", True},
            "callback_requested": callback in {"yes", True},
            "summary": self._summary(answers),
            "transcript": session.transcript_text,
            "structured_response": answers,
            "duration": session.duration_seconds,
            "timestamp": (session.ended_at or session.started_at).isoformat(),
        }

    def _summary(self, answers: dict[str, Any]) -> str:
        return "; ".join(f"{k}: {v}" for k, v in answers.items() if v not in (None, ""))

    def _flat_call(self, call: dict[str, Any]) -> dict[str, Any]:
        row = {key: call.get(key, "") for key in EXPORT_HEADERS if key != "structured_response"}
        row["structured_response"] = json.dumps(call.get("structured_response", {}), ensure_ascii=False)
        return row

    def _clean_text(self, value: Any) -> str:
        return re.sub(r"\s+", " ", str(value or "")).strip()

    def _normalize_phone(self, value: Any) -> str:
        raw = str(value or "").strip()
        if not raw:
            return ""
        leading_plus = raw.startswith("+")
        cleaned = re.sub(r"[^\d]", "", raw)
        return f"+{cleaned}" if leading_plus and cleaned else cleaned

    def _phone_key(self, value: Any) -> str:
        return re.sub(r"\D", "", str(value or ""))
