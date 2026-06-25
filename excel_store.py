from __future__ import annotations

from pathlib import Path
from threading import Lock

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:  # pragma: no cover - exercised only when dependency is missing
    Workbook = None
    load_workbook = None
    Worksheet = object

from models import CallSession
from prompts import ANSWER_FIELDS

SHEET_NAME = "Call Answers"
HEADERS = [
    "call_id",
    "started_at",
    "ended_at",
    "duration_seconds",
    "status",
    *(field.name for field in ANSWER_FIELDS),
    "transcript",
]


class ExcelAnswerStore:
    """Thread-safe append-only Excel store for captured call answers.

    The store intentionally keeps Excel as a reporting/export layer. Its public
    API accepts the domain-level ``CallSession`` object so future GSM/Asterisk
    adapters can persist results without depending on browser-specific code.
    """

    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path
        self._lock = Lock()

    def append_call(self, session: CallSession) -> None:
        if Workbook is None or load_workbook is None:
            raise RuntimeError("openpyxl is required to save call answers. Install requirements.txt.")
        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
        with self._lock:
            workbook = self._load_or_create_workbook()
            sheet = self._get_answer_sheet(workbook)
            sheet.append(self._session_to_row(session))
            self._format_sheet(sheet)
            workbook.save(self.workbook_path)
            workbook.close()

    def _load_or_create_workbook(self):
        if self.workbook_path.exists():
            workbook = load_workbook(self.workbook_path)
            sheet = self._get_answer_sheet(workbook)
            self._ensure_headers(sheet)
            return workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAME
        sheet.append(HEADERS)
        self._format_sheet(sheet)
        return workbook

    def _get_answer_sheet(self, workbook) -> Worksheet:
        if SHEET_NAME in workbook.sheetnames:
            return workbook[SHEET_NAME]
        sheet = workbook.active
        if sheet.max_row == 1 and all(cell.value is None for cell in sheet[1]):
            sheet.title = SHEET_NAME
        else:
            sheet = workbook.create_sheet(SHEET_NAME)
        self._ensure_headers(sheet)
        return sheet

    def _ensure_headers(self, sheet: Worksheet) -> None:
        existing = [cell.value for cell in sheet[1]] if sheet.max_row else []
        if existing[: len(HEADERS)] == HEADERS:
            return
        if sheet.max_row == 1 and all(value is None for value in existing):
            sheet.delete_rows(1)
        sheet.insert_rows(1)
        for column, header in enumerate(HEADERS, start=1):
            sheet.cell(row=1, column=column, value=header)

    def _session_to_row(self, session: CallSession) -> list[object]:
        return [
            session.call_id,
            session.started_at.isoformat(),
            session.ended_at.isoformat() if session.ended_at else "",
            session.duration_seconds,
            session.status,
            *(session.answers.get(field.name, "unknown") for field in ANSWER_FIELDS),
            session.transcript_text,
        ]

    def _format_sheet(self, sheet: Worksheet) -> None:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.style = "Headline 3"
        widths = {
            "A": 38,
            "B": 28,
            "C": 28,
            "D": 18,
            "E": 20,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        for column in range(6, len(HEADERS) + 1):
            letter = sheet.cell(row=1, column=column).column_letter
            sheet.column_dimensions[letter].width = 26 if column < len(HEADERS) else 90

